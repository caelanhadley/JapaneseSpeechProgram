from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import sys

# Instance Variables #
PATH = "answer_sheets\\"
NUMBER_OF_TAGS = 9 # CONFIGURES RETRIVE ANSWERS METHOD
name = "L4AS.xlsx"

# Init #
answer_book = Workbook()
answer_book = load_workbook(PATH + name)
answer_sheet = answer_book.active






# DATA FORMAT:                                                           #
# [1 DESCRIPTOR][2 TAG][3 CATEGORY][4 LESSON][5 Image?/Type][6 Hint][7 Interactions Text][8 Override][9 Reference][ANSWER][ANSWER][ANSWER]... #

# Retrives all relavent answers to a question (id)
# Excludes row 1, because row 1 is reserved for id descriptors


def retriveAnswers(id):
    result = []
    # Note the '- 1' accounts for the descriptor column not being included
    for i in (range(answer_sheet.max_column - NUMBER_OF_TAGS)):
        if answer_sheet.cell(row=id, column=i + NUMBER_OF_TAGS + 1).value != "" and answer_sheet.cell(row=id, column=i + NUMBER_OF_TAGS + 1).value != None:
            result.append(answer_sheet.cell(row=id, column=i + NUMBER_OF_TAGS + 1).value)
    return result

# Determines the answer format; X - Exact; V - Variable; X > Compare Answers; V > Compare Contains Asnwer
def getTag(id):
    return answer_sheet.cell(row=id, column=2).value


def getCategory(id):
    return answer_sheet.cell(row=id, column=3).value


def getLesson(id):
    return answer_sheet.cell(row=id, column=4).value


def getImage(id):
    return answer_sheet.cell(row=id, column=5).value


def getHint(id):
    return answer_sheet.cell(row=id, column=6).value

def getInteraction(id):
    return answer_sheet.cell(row=id, column=7).value

def getAltId(id):
    return answer_sheet.cell(row=id, column=8).value

def hasImage(id):
    if answer_sheet.cell(row=id, column=5).value != None and answer_sheet.cell(row=id, column=5).value != '':
        return True
    else:
        return False

def hasInteractions(id):
    if answer_sheet.cell(row=id, column=7).value != None and answer_sheet.cell(row=id, column=7).value != '':
        return True
    else:
        return False

def hasOverride(id): # Overrides the audio file to be played, useful for questions with same audio
    if answer_sheet.cell(row=id, column=8).value != None and answer_sheet.cell(row=id, column=8).value != '':
        return True
    else:
        return False
    
def hasReference(id):
    if answer_sheet.cell(row=id, column=9).value != None and answer_sheet.cell(row=id, column=9).value != '':
        return True
    else:
        return False

# Descriptor, a string representing a description
# Returns an array of matching answer's id


def findAnswersFromDesc(descriptor):
    match = []
    for i in range(getRows()):
        if descriptor == answer_sheet.cell(row=i + 1, column=1).value:
            match.append((i + 1))
    return match


def getRows():
    number_of_cells = 0
    for i in range(answer_sheet.max_row):
        if answer_sheet.cell(i + 1, 1).value != None:
            number_of_cells += 1
    return number_of_cells


def compareAnswers(answerIn, id):
    answers = retriveAnswers(id)
    for i in range(len(answers)):
        if answers[i] == answerIn:
            return True
    return False

# Def compareContains(answer, id) Checks if an answer contains @param 'answer'


def compareContains(answerIn, id):
    answers = retriveAnswers(id)
    for i in range(len(answers)):
        if answers[i] in answerIn:
            return True
    return False


def getQuestionSet(lessons):
    result = []
    for i in range(len(lessons)):
        for b in range(getRows()):
            if lessons[i] == answer_sheet.cell(row=b + 1, column=4).value:
                result.append(b + 1)
    return result
